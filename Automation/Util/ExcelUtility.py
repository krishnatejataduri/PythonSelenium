'''
@author: Krishna Teja
@date: 11/12/2019
'''

from openpyxl import *


class ExcelUtility:

    def __init__(self, path, sheet_name):
        self.path = path
        self.sheet_name = sheet_name
        self.ws = self._open_excel()
        self.cell = None
        self.column_dictionary = {}
        self._read_columns()

    def _open_excel(self):
        '''
        This method is called in the __init__ method. Not be accessed outside of __init__.
        Opens an excel workbook and returns the worksheet object.
        '''
        return load_workbook(self.path).get_sheet_by_name(self.sheet_name)

    def _read_columns(self):
        '''
        This method is called in the __init__ method. Not be accessed outside of __init__.
        Reads the column headers and stores indices in a dictionary object declared inside __init__.
        '''
        for col in range(1, self.ws.max_column + 1):
            self.column_dictionary[self.ws.cell(1, col).value] = col

    def get_cell_value(self, column_name, row_num):
        '''
        Returns the value in a specific cell by taking the Column name and row number.
        '''
        return str(self.ws.cell(row_num, self.column_dictionary[column_name]).value)

    def get_row_count(self):
        '''
        Returns the total number of rows
        '''
        return self.ws.max_row

    def get_test_list(self):
        '''
        Returns the total number of test cases to be executed from Runner file
        '''
        current_row = 1
        test_list = []
        while current_row <= self.get_row_count():
            if self.get_cell_value("ExecutionFlag", current_row) == 'Y':
                test_details = {'TestcaseName': self.get_cell_value("TestcaseName", current_row),
                                'DebugMode': self.get_cell_value("DebugMode", current_row),
                                'DatasetName': self.get_cell_value("DatasetName", current_row),
                                'Browser': self.get_cell_value("Browser", current_row),
                                'PriorTest': self.get_cell_value("PriorTest", current_row)}
                test_list.append(test_details)
            current_row += 1
        return test_list

    def get_config_info(self):
        '''
        Returns the execution configuration info from the Runner file
        '''
        current_row = 2
        config_details = {'TestSetName': self.get_cell_value("TestSetName", current_row),
                          'Environment': self.get_cell_value("Environment", current_row),
                          'Parallel': self.get_cell_value("Parallel", current_row),
                          'Instances': self.get_cell_value("Instances", current_row),
                          'Remote': self.get_cell_value("Remote", current_row)}

        return config_details

    def get_test_steps(self,execution_environment):
        '''
        Returns the Test step details for all the steps in a test case.
        '''
        current_row = 2
        test_step_list = []
        while current_row <= self.get_row_count():
            #Condition to consider the step: Exec Flag should be Y and the current Env should match the step Env or ste Env should be All.
            if self.get_cell_value("ExecutionFlag", current_row) == 'Y' and self.get_cell_value("Environment", current_row).strip() in [execution_environment , 'All'] :
                step_entry = {'StepID': self.get_cell_value("StepID", current_row),
                              'ScreenName': self.get_cell_value("ScreenName", current_row),
                              'ObjectDescription': self.get_cell_value("ObjectDescription", current_row),
                              'ObjectName': self.get_cell_value("ObjectName", current_row),
                              'Action': self.get_cell_value("Action", current_row),
                              'Function': self.get_cell_value("Function", current_row),
                              'UserInput': self.get_cell_value("UserInput", current_row),
                              'Environment': self.get_cell_value("Environment", current_row)}
                test_step_list.append(step_entry)
            current_row += 1
        return test_step_list

    def load_object_repo(self):
        '''Returns a list of dictionaries with object properties of a particular screen'''
        current_row = 2
        object_repo_dict = {}
        while current_row <= self.get_row_count():
            object_record_dict = {'ObjectDescription': self.get_cell_value('ObjectDescription', current_row),
                                  'ObjectType': self.get_cell_value('ObjectType', current_row),
                                  'LocatorType': self.get_cell_value('LocatorType', current_row),
                                  'LocatorValue': self.get_cell_value('LocatorValue', current_row)}
            object_repo_dict.update({self.get_cell_value('ObjectName', current_row): object_record_dict})
            current_row+=1
        return object_repo_dict
